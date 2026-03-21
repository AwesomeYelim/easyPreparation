"""
스프레드시트(xlsx) 생성 스크립트 템플릿
usage: python generate_sheet.py <출력경로> [--data <json_file>]

생성된 .xlsx 파일은 drive_uploadFile MCP tool로 업로드하거나
upload_sheet.py 로 Google Sheets 변환 업로드 가능.
"""

import sys
import os
import argparse
import json
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ── 스타일 헬퍼 ────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def header_cell(ws, row: int, col: int, value,
                bg="1A237E", fg="FFFFFF", size=10, bold=True, center=False):
    """헤더 스타일 셀 작성"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=size, color=fg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    return cell


def data_cell(ws, row: int, col: int, value, bg=None, bold=False):
    """데이터 셀 작성"""
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return cell


def title_row(ws, row: int, col_span: int, value,
              bg="1A237E", fg="FFFFFF", size=14):
    """병합 타이틀 행 작성"""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=col_span,
    )
    cell = ws.cell(row=row, column=1, value=value)
    cell.fill = fill(bg)
    cell.font = Font(bold=True, size=size, color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    return cell


def section_header(ws, row: int, value, bg="3949AB", fg="FFFFFF"):
    """섹션 구분 헤더"""
    cell = ws.cell(row=row, column=1, value=value)
    cell.fill = fill(bg)
    cell.font = Font(bold=True, size=11, color=fg)
    return cell


# ── 시트 빌더 ─────────────────────────────────────────────

def _build_overview(wb, data: dict):
    ws = wb.create_sheet("개요")
    title_row(ws, 1, 3, data.get("title", "분석 보고서"), bg="1A237E", size=16)
    title_row(ws, 2, 3, data.get("subtitle", ""), bg="E8EAF6", fg="000000", size=10)

    section_header(ws, 4, "■ 모듈 기본 정보", bg="3949AB")
    for i, h in enumerate(["항목", "값", "비고"], 1):
        header_cell(ws, 5, i, h, bg="5C6BC0")
    for r, row in enumerate(data.get("info_rows", []), 6):
        for c, val in enumerate(row, 1):
            data_cell(ws, r, c, val)

    base_row = 6 + len(data.get("info_rows", [])) + 2
    section_header(ws, base_row, "■ 패키지 구조", bg="3949AB")
    for i, h in enumerate(["패키지", "주요 파일", "역할"], 1):
        header_cell(ws, base_row + 1, i, h, bg="5C6BC0")
    for r, row in enumerate(data.get("package_rows", []), base_row + 2):
        for c, val in enumerate(row, 1):
            data_cell(ws, r, c, val)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45


def _build_startup(wb, data: dict):
    ws = wb.create_sheet("기동 흐름")
    title_row(ws, 1, 4, data.get("title", "기동 흐름"), bg="1B5E20", size=14)

    for i, h in enumerate(["순서", "함수", "패키지", "설명"], 1):
        header_cell(ws, 3, i, h, bg="2E7D32")
    for r, row in enumerate(data.get("steps", []), 4):
        for c, val in enumerate(row, 1):
            data_cell(ws, r, c, val)

    base = 4 + len(data.get("steps", [])) + 2
    section_header(ws, base, "■ 생성되는 고루틴", bg="1B5E20")
    for i, h in enumerate(["고루틴", "기동 위치", "역할"], 1):
        header_cell(ws, base + 1, i, h, bg="388E3C")
    for r, row in enumerate(data.get("routines", []), base + 2):
        for c, val in enumerate(row, 1):
            data_cell(ws, r, c, val)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 55


def _build_functions(wb, data: dict):
    pkg_colors = data.get("pkg_colors", {})
    funcs = data.get("functions", [])

    ws = wb.create_sheet("함수 목록")
    title_row(ws, 1, 4, f"전체 함수 목록 ({len(funcs)}개)", bg="4A148C", size=14)

    for i, h in enumerate(["패키지", "함수명", "분류", "설명"], 1):
        header_cell(ws, 3, i, h, bg="6A1B9A")
    for r, (pkg, fn, cat, desc) in enumerate(funcs, 4):
        bg = pkg_colors.get(pkg, "FFFFFF")
        data_cell(ws, r, 1, pkg, bg=bg)
        data_cell(ws, r, 2, fn,  bg=bg, bold=True)
        data_cell(ws, r, 3, cat, bg=bg)
        data_cell(ws, r, 4, desc, bg=bg)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50


def _build_flow(wb, data: dict):
    ws = wb.create_sheet("처리 흐름")
    title_row(ws, 1, 4, data.get("title", "처리 흐름"), bg="BF360C", size=14)

    row = 3
    for section in data.get("sections", []):
        section_header(ws, row, section["name"], bg="BF360C")
        row += 1
        for i, h in enumerate(section["headers"], 1):
            header_cell(ws, row, i, h, bg="D84315")
        row += 1
        for r_data in section["rows"]:
            for c, val in enumerate(r_data, 1):
                data_cell(ws, row, c, val)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 45


def _build_deps(wb, data: dict):
    ws = wb.create_sheet("의존성")
    title_row(ws, 1, 5, data.get("title", "의존성"), bg="006064", size=14)

    row = 3
    for section in data.get("sections", []):
        section_header(ws, row, section["name"], bg="006064")
        row += 1
        for i, h in enumerate(section["headers"], 1):
            header_cell(ws, row, i, h, bg="00838F")
        row += 1
        for r_data in section["rows"]:
            for c, val in enumerate(r_data, 1):
                data_cell(ws, row, c, val)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 40


# ── build() ───────────────────────────────────────────────

def build(output_path: str, data: dict) -> str:
    wb = openpyxl.Workbook()
    del wb["Sheet"]  # 기본 시트 제거

    # ── 여기서부터 시트 내용 추가 ──────────────────────────────
    _build_overview(wb, data.get("overview", {}))
    _build_startup(wb, data.get("startup", {}))
    _build_functions(wb, data.get("functions", {}))
    _build_flow(wb, data.get("flow", {}))
    _build_deps(wb, data.get("deps", {}))
    # ── 시트 내용 끝 ──────────────────────────────────────────

    wb.save(output_path)
    return output_path


# ── 기본 데이터 (SA 모듈 분석) ─────────────────────────────

def _default_data() -> dict:
    return {
        "overview": {
            "title": "SA 모듈 분석 보고서",
            "subtitle": "생성일: 2026-03-05  |  분석 대상: cmd/sa/  |  함수: 187개  |  파일: 33개  |  라인: 4,438줄",
            "info_rows": [
                ("모듈명", "SA (System Agent)", "에이전트 시스템의 중앙 오케스트레이터"),
                ("엔트리 포인트", "cmd/sa/sa.go", "main() 함수 위치"),
                ("총 파일 수", 33, "플랫폼 공통 31 + linux/windows 2"),
                ("총 함수 수", 187, ""),
                ("총 라인 수", 4438, ""),
                ("예상 토큰", 17752, "indexer 기준"),
                ("플랫폼 전용 함수", "InitProcessCheck, InitRateLimit", "utils/ (linux/windows)"),
            ],
            "package_rows": [
                ("cmd/sa/", "sa.go", "main() - 전체 초기화 및 루틴 기동"),
                ("moduleManager/", "moduleManager.go, procModuleRoutine.go, messageRoutine.go", "모듈 생명주기 관리 (기동/종료/업데이트/상태)"),
                ("saJobManager/", "saJobManager.go", "SA Job 큐 처리 (init/stop/kill/configUpdate 등)"),
                ("messageSender/", "messageSender.go, jobRequestRoutine.go", "AM 서버로 메시지/로그/결과 전송 (HTTP)"),
                ("logCollectionManager/", "logCollectionManager.go", "각 모듈 로그 파일 수집 및 관리"),
                ("resourceManager/", "resourceManager.go", "CPU/MEM/DISK/Agent 리소스 모니터링 및 알람"),
                ("grpcServer/", "server.go, connect.go", "gRPC 서버 (모듈 → SA 통신 수신)"),
                ("dbManager/", "dbManager.go, filedb.go, job.go, upgradeJob.go", "BoltDB 기반 로컬 Job 영속화"),
                ("define/", "define.go, messageSenderDefine.go", "구조체/상수 정의 (StRunMem 등)"),
                ("global/", "channel.go, message.go", "전역 채널 및 메시지 타입 정의"),
                ("utils/", "utils_linux.go, utils_windows.go", "플랫폼별 프로세스 체크/레이트리밋"),
            ],
        },
        "startup": {
            "title": "SA 기동 흐름 (main → 루틴 기동)",
            "steps": [
                (1,  "config.ExecuteOption",                         "internal/application/config",        "CLI 옵션 파싱"),
                (2,  "config.InitProcess",                           "internal/application/config",        "설정 파일 로드, 로그 초기화"),
                (3,  "utils.InitProcessCheck",                       "cmd/sa/utils/",                      "중복 프로세스 확인 (플랫폼별)"),
                (4,  "utils.InitRateLimit",                          "cmd/sa/utils/",                      "재시작 레이트 리밋 초기화"),
                (5,  "define.StRunMem.InitializeRunMem",             "cmd/sa/define/",                     "런타임 메모리 초기화"),
                (6,  "stModuleState.InitializeModuleState",          "cmd/sa/moduleManager/",              "모듈 상태맵 초기화"),
                (7,  "dbManager.InitDbManager",                      "cmd/sa/dbManager/",                  "BoltDB 오픈 → RoutineDbManager 고루틴 기동"),
                (8,  "messageSender.InitMessageSender",              "cmd/sa/messageSender/",              "RoutineMessageSender 고루틴 기동"),
                (9,  "messageSender.InitSenderProcess",              "cmd/sa/messageSender/",              "HTTP 클라이언트 초기화"),
                (10, "grpcServer.StartGrpcServer",                   "cmd/sa/grpcServer/",                 "gRPC 리스너 바인딩 및 Serve"),
                (11, "moduleManager.InitModuleManager",              "cmd/sa/moduleManager/",              "RoutineProcModule, RoutineJobSave, RoutineModuleMessage, RoutineGatherService 기동"),
                (12, "resourceManager.RoutineResourceManager",       "cmd/sa/resourceManager/",            "CPU/MEM/DISK/Agent 모니터 고루틴 4개 기동"),
                (13, "logCollectionManager.RoutineLogCollectionManager", "cmd/sa/logCollectionManager/",   "로그 수집/관리 고루틴 기동"),
                (14, "saJobManager.RoutineProcSaJobManager",         "cmd/sa/saJobManager/",               "SA Job 처리 루프 기동"),
                (15, "wgSa.Wait",                                    "cmd/sa/sa.go",                       "전체 고루틴 종료 대기"),
            ],
            "routines": [
                ("RoutineDbManager",           "InitDbManager",                  "BoltDB Job 읽기/쓰기"),
                ("RoutineMessageSender",       "InitMessageSender",              "AM 서버 메시지 전송"),
                ("RoutineModuleMessage",       "InitModuleManager",              "모듈 gRPC 메시지 수신"),
                ("RoutineProcModule",          "InitModuleManager",              "모듈 기동/종료/업데이트 메인 루프"),
                ("RoutineJobSave",             "InitModuleManager",              "Job 저장"),
                ("RoutineGatherService",       "InitModuleManager",              "GS 스케줄 관리"),
                ("routineResourceMonitor×4",   "RoutineResourceManager",         "CPU/MEM/DISK/Agent 리소스 모니터"),
                ("routineLogCollection",       "RoutineLogCollectionManager",    "모듈 로그 파일 읽기"),
                ("routineLogManagement",       "RoutineLogCollectionManager",    "로그 파일 정리/롤링"),
                ("RoutineProcSaJobManager",    "main",                           "SA Job 명령 처리 루프"),
            ],
        },
        "functions": {
            "pkg_colors": {
                "sa": "EDE7F6", "moduleManager": "E3F2FD", "saJobManager": "E8F5E9",
                "messageSender": "FFF8E1", "grpcServer": "FCE4EC", "dbManager": "F3E5F5",
                "logCollectionManager": "E0F2F1", "resourceManager": "FBE9E7",
                "define": "F5F5F5", "global": "FAFAFA", "utils": "E8EAF6",
            },
            "functions": [
                ("sa",                    "main",                        "진입점",    "SA 프로세스 시작점, 모든 루틴 초기화"),
                ("moduleManager",         "InitModuleManager",           "초기화",    "4개 핵심 루틴 기동"),
                ("moduleManager",         "RoutineProcModule",           "루틴",      "모듈 기동/종료/업데이트 메인 루프"),
                ("moduleManager",         "RoutineModuleMessage",        "루틴",      "모듈 gRPC 메시지 수신 처리"),
                ("moduleManager",         "RoutineJobSave",              "루틴",      "Job 저장 처리"),
                ("moduleManager",         "RoutineGatherService",        "루틴",      "GS 스케줄 관리"),
                ("moduleManager",         "runModule",                   "모듈실행",  "Local/Container 방식 선택 후 실행"),
                ("moduleManager",         "runModuleInLocal",            "모듈실행",  "로컬 프로세스로 모듈 실행"),
                ("moduleManager",         "runModuleInContainer",        "모듈실행",  "컨테이너 환경에서 모듈 실행"),
                ("moduleManager",         "runCceModuleInLocal",         "모듈실행",  "CCE 모듈 로컬 실행"),
                ("moduleManager",         "runCveModuleInLocal",         "모듈실행",  "CVE 모듈 로컬 실행"),
                ("moduleManager",         "runRunner",                   "모듈실행",  "Runner 프로세스 실행"),
                ("moduleManager",         "runMetieyeStart",             "모듈실행",  "Metieye 모듈 시작"),
                ("moduleManager",         "runUpdateJob",                "업데이트",  "업데이트 Job 실행"),
                ("moduleManager",         "prepareRunner",               "모듈실행",  "Runner 실행 전 상태 등록"),
                ("moduleManager",         "cleanupRunner",               "정리",      "Runner 정리"),
                ("moduleManager",         "cleanupTempDirectory",        "정리",      "임시 디렉토리 정리"),
                ("moduleManager",         "cleanup",                     "정리",      "모듈 리소스 정리"),
                ("moduleManager",         "cleanupModule",               "정리",      "특정 모듈 정리"),
                ("moduleManager",         "replaceModule",               "업데이트",  "모듈 바이너리 교체"),
                ("moduleManager",         "updateModule",                "업데이트",  "모듈 설정 업데이트"),
                ("moduleManager",         "updateAm",                    "업데이트",  "AM 설정 업데이트"),
                ("moduleManager",         "updateSa",                    "업데이트",  "SA 설정 업데이트"),
                ("moduleManager",         "updateAgent",                 "업데이트",  "에이전트 전체 설정 갱신"),
                ("moduleManager",         "updateConfigMem",             "업데이트",  "설정 메모리 갱신"),
                ("moduleManager",         "checkModuleRun",              "상태확인",  "모듈 실행 여부 체크"),
                ("moduleManager",         "checkUpdataInit",             "상태확인",  "업데이트 초기화 체크"),
                ("moduleManager",         "isModuleUpdate",              "상태확인",  "모듈 업데이트 필요 여부"),
                ("moduleManager",         "isAmUpdate",                  "상태확인",  "AM 업데이트 필요 여부"),
                ("moduleManager",         "isCmUpdate",                  "상태확인",  "CM 업데이트 필요 여부"),
                ("moduleManager",         "getModuleState",              "상태",      "모듈 상태 조회"),
                ("moduleManager",         "moduleStateUpdate",           "상태",      "모듈 상태 갱신"),
                ("moduleManager",         "moduleUpdateStateAll",        "상태",      "전체 모듈 업데이트 상태 설정"),
                ("moduleManager",         "ModuleStateClear",            "상태",      "모듈 상태 초기화"),
                ("moduleManager",         "AddModuleState",              "상태",      "모듈 상태 추가"),
                ("moduleManager",         "getAllJobData",               "Job",       "전체 Job 데이터 조회"),
                ("moduleManager",         "searchJobModule",             "Job",       "Job 모듈 검색"),
                ("moduleManager",         "sendSaJob",                   "전송",      "SA Job 전송"),
                ("moduleManager",         "sendGsJob",                   "전송",      "GS Job 전송"),
                ("moduleManager",         "sendJobDone",                 "전송",      "Job 완료 전송"),
                ("moduleManager",         "sendJobFail",                 "전송",      "Job 실패 전송"),
                ("moduleManager",         "sendJobSuccess",              "전송",      "Job 성공 전송"),
                ("moduleManager",         "sendAfInfo",                  "전송",      "AF 정보 전송"),
                ("moduleManager",         "sendGsInfo",                  "전송",      "GS 정보 전송"),
                ("moduleManager",         "sendFinish",                  "전송",      "완료 신호 전송"),
                ("moduleManager",         "sendModuleStatusLog",         "전송",      "모듈 상태 로그 전송"),
                ("moduleManager",         "sendResourceAlarm",           "전송",      "리소스 알람 전송"),
                ("moduleManager",         "sendAgentLog",                "전송",      "에이전트 로그 전송"),
                ("moduleManager",         "cvtStrAlarmTarget",           "유틸",      "알람 대상 문자열 변환"),
                ("moduleManager",         "procModuleInit",              "처리",      "모듈 초기화 처리"),
                ("moduleManager",         "procAmModuleInit",            "처리",      "AM 모듈 초기화 처리"),
                ("moduleManager",         "procModuleFinish",            "처리",      "모듈 종료 처리"),
                ("moduleManager",         "procModuleDownCase",          "처리",      "모듈 다운 케이스 처리"),
                ("moduleManager",         "procRunnerFinish",            "처리",      "Runner 종료 처리"),
                ("moduleManager",         "procTerminateModule",         "처리",      "모듈 강제 종료 처리"),
                ("moduleManager",         "procModuleKill",              "처리",      "모듈 Kill 처리"),
                ("moduleManager",         "procStopAll",                 "처리",      "전체 모듈 정지"),
                ("moduleManager",         "procDbCheck",                 "처리",      "DB 연결 확인 처리"),
                ("moduleManager",         "procSaRestart",               "처리",      "SA 재시작 처리"),
                ("moduleManager",         "subRoutineProcMessage",       "서브루틴",  "메시지 처리 서브루틴"),
                ("moduleManager",         "subRoutineProcLogCollection", "서브루틴",  "로그 수집 서브루틴"),
                ("moduleManager",         "handleOriginLogPath",         "로그",      "원본 로그 경로 처리"),
                ("moduleManager",         "handleTempLogPath",           "로그",      "임시 로그 경로 처리"),
                ("moduleManager",         "handleTarLogPath",            "로그",      "TAR 로그 경로 처리"),
                ("moduleManager",         "createTarGzFile",             "로그",      "TAR.GZ 파일 생성"),
                ("moduleManager",         "setInitModule",               "초기화",    "모듈 초기화 설정"),
                ("moduleManager",         "setModuleConfig",             "설정",      "모듈 설정 적용"),
                ("moduleManager",         "setAmConfig",                 "설정",      "AM 설정 적용"),
                ("moduleManager",         "setSaConfig",                 "설정",      "SA 설정 적용"),
                ("saJobManager",          "RoutineProcSaJobManager",     "루틴",      "SA Job 처리 메인 루프"),
                ("saJobManager",          "subRoutineProcSaJob",         "서브루틴",  "SA Job 타입별 처리 분기"),
                ("saJobManager",          "procInit",                    "처리",      "초기화 Job (설정 로드, 모듈 기동)"),
                ("saJobManager",          "procStop",                    "처리",      "모듈 정지 Job 처리"),
                ("saJobManager",          "procOff",                     "처리",      "모듈 비활성화 처리"),
                ("saJobManager",          "procConfigUpdate",            "처리",      "설정 업데이트 처리"),
                ("saJobManager",          "procAgentLog",                "처리",      "에이전트 로그 처리"),
                ("saJobManager",          "stopService",                 "처리",      "서비스 정지"),
                ("saJobManager",          "enableSendInit",              "상태",      "전송 초기화 활성화"),
                ("saJobManager",          "disableSendInit",             "상태",      "전송 초기화 비활성화"),
                ("saJobManager",          "AllModuleStop",               "제어",      "전체 모듈 정지"),
                ("messageSender",         "InitMessageSender",           "초기화",    "메시지 발신 루틴 기동"),
                ("messageSender",         "InitSenderProcess",           "초기화",    "HTTP 클라이언트 초기화"),
                ("messageSender",         "RoutineMessageSender",        "루틴",      "메시지 발신 루프"),
                ("messageSender",         "RoutineJobRequest",           "루틴",      "Job 요청 루틴 (AM 폴링)"),
                ("messageSender",         "sendPostRequest",             "전송",      "HTTP POST 요청"),
                ("messageSender",         "sendPostRequestWithRetry",    "전송",      "재시도 포함 HTTP POST"),
                ("messageSender",         "sendDeleteRequest",           "전송",      "HTTP DELETE 요청"),
                ("messageSender",         "SendResultData",              "전송",      "결과 데이터 전송"),
                ("messageSender",         "SendLogData",                 "전송",      "로그 데이터 전송"),
                ("messageSender",         "SendDbCheckData",             "전송",      "DB 확인 데이터 전송"),
                ("messageSender",         "SendAgentLogMessage",         "전송",      "에이전트 로그 메시지 전송"),
                ("messageSender",         "CommonProcess",               "공통",      "공통 gRPC 처리"),
                ("messageSender",         "getJobResponse",              "Job",       "Job 응답 처리"),
                ("messageSender",         "handleJobResponse",           "Job",       "Job 응답 핸들러"),
                ("messageSender",         "getUpgradeJobResponse",       "Job",       "업그레이드 Job 응답 처리"),
                ("grpcServer",            "StartGrpcServer",             "초기화",    "gRPC 서버 시작"),
                ("grpcServer",            "NewGrpcServer",               "초기화",    "gRPC 서버 인스턴스 생성"),
                ("grpcServer",            "HeartBeat",                   "gRPC",      "모듈 → SA 하트비트 수신"),
                ("grpcServer",            "Accept",                      "gRPC",      "연결 수락"),
                ("grpcServer",            "Close",                       "gRPC",      "연결 종료"),
                ("grpcServer",            "checkCloseConnection",        "gRPC",      "연결 종료 여부 확인"),
                ("grpcServer",            "sendGrpcClose",               "gRPC",      "gRPC 종료 신호 전송"),
                ("grpcServer",            "GetGrpcInfo",                 "조회",      "gRPC 연결 정보 조회"),
                ("grpcServer",            "SetGrpcInfo",                 "설정",      "gRPC 연결 정보 설정"),
                ("grpcServer",            "DeleteGrpcInfo",              "관리",      "gRPC 연결 정보 삭제"),
                ("grpcServer",            "NewGrpcConnInfo",             "생성",      "gRPC 연결 정보 생성"),
                ("dbManager",             "InitDbManager",               "초기화",    "DB 매니저 초기화 및 루틴 기동"),
                ("dbManager",             "RoutineDbManager",            "루틴",      "DB 처리 메인 루틴"),
                ("dbManager",             "OpenFileDb",                  "DB",        "파일 DB 오픈"),
                ("dbManager",             "CloseFileDb",                 "DB",        "파일 DB 닫기"),
                ("dbManager",             "InsertJob",                   "Job",       "Job 삽입"),
                ("dbManager",             "UpdateJob",                   "Job",       "Job 업데이트"),
                ("dbManager",             "DeleteJob",                   "Job",       "Job 삭제"),
                ("dbManager",             "GetJob",                      "Job",       "Job 조회"),
                ("dbManager",             "GetJobByName",                "Job",       "이름으로 Job 조회"),
                ("dbManager",             "SelectAllJob",                "Job",       "전체 Job 조회"),
                ("dbManager",             "UpdateJobRunFlag",            "Job",       "Job 실행 플래그 업데이트"),
                ("dbManager",             "InsertUpgradeJob",            "UpgradeJob","업그레이드 Job 삽입"),
                ("dbManager",             "DeleteUpgradeJob",            "UpgradeJob","업그레이드 Job 삭제"),
                ("dbManager",             "GetAllUpgradeJob",            "UpgradeJob","전체 업그레이드 Job 조회"),
                ("logCollectionManager",  "RoutineLogCollectionManager", "루틴",      "로그 수집 매니저 루틴"),
                ("logCollectionManager",  "routineLogCollection",        "루틴",      "로그 수집 루틴"),
                ("logCollectionManager",  "routineLogManagement",        "루틴",      "로그 관리 루틴"),
                ("logCollectionManager",  "collectLog",                  "수집",      "로그 파일 읽기 및 수집"),
                ("logCollectionManager",  "collectAgentLog",             "수집",      "에이전트 로그 수집"),
                ("logCollectionManager",  "collectCpuUsage",             "수집",      "CPU 사용량 수집"),
                ("logCollectionManager",  "startLogCollection",          "상태",      "로그 수집 시작"),
                ("logCollectionManager",  "endLogCollection",            "상태",      "로그 수집 종료"),
                ("logCollectionManager",  "logCollectionAlarm",          "알람",      "로그 수집 알람"),
                ("logCollectionManager",  "GetLogPath",                  "조회",      "로그 경로 조회"),
                ("logCollectionManager",  "InitializeLogState",          "초기화",    "로그 상태 초기화"),
                ("resourceManager",       "RoutineResourceManager",      "루틴",      "리소스 모니터 루틴 4개 기동"),
                ("resourceManager",       "routineResourceMonitorAgent", "모니터",    "에이전트 리소스 모니터"),
                ("resourceManager",       "routineResourceMonitorCpu",   "모니터",    "CPU 사용률 모니터"),
                ("resourceManager",       "routineResourceMonitorMem",   "모니터",    "메모리 사용률 모니터"),
                ("resourceManager",       "routineResourceMonitorDisk",  "모니터",    "디스크 사용률 모니터"),
                ("define",                "InitializeRunMem",            "초기화",    "런타임 메모리 전체 초기화"),
                ("global",                "SendMessageToChannel",        "채널",      "채널로 메시지 전송"),
                ("global",                "ReceiveMessageFromChannel",   "채널",      "채널에서 메시지 수신"),
                ("global",                "SendAgentLogMessage",         "채널",      "에이전트 로그 메시지 전송"),
                ("utils",                 "InitProcessCheck",            "초기화",    "프로세스 중복 확인 초기화"),
                ("utils",                 "InitRateLimit",               "초기화",    "레이트 리밋 초기화"),
            ],
        },
        "flow": {
            "title": "SA 핵심 처리 흐름",
            "sections": [
                {
                    "name": "■ Job 처리 흐름 (AM → SA → 모듈)",
                    "headers": ["단계", "컴포넌트", "함수", "설명"],
                    "rows": [
                        ("1. Job 수신",  "messageSender",     "RoutineJobRequest → getJobResponse",             "AM 서버에서 HTTP로 Job 폴링"),
                        ("2. Job 저장",  "dbManager",         "InsertJob → BoltDB",                             "수신 Job을 로컬 DB에 영속화"),
                        ("3. Job 분배",  "RoutineProcModule", "getAllJobData → searchJobModule",                 "DB에서 Job 읽어 대상 모듈 결정"),
                        ("4. SA Job",    "saJobManager",      "RoutineProcSaJobManager → subRoutineProcSaJob",  "SA 자체 처리 Job (init/stop/configUpdate)"),
                        ("5. 모듈 실행", "moduleManager",     "runModule → runModuleInLocal/Container",          "해당 모듈 바이너리 프로세스 기동"),
                        ("6. 결과 수신", "grpcServer",        "HeartBeat / CommonProcess",                      "모듈 → SA gRPC로 상태/완료 보고"),
                        ("7. 결과 전송", "messageSender",     "SendResultData → sendPostRequest",               "AM 서버로 HTTP POST 결과 전송"),
                        ("8. Job 삭제",  "dbManager",         "DeleteJob",                                      "처리 완료된 Job DB에서 제거"),
                    ],
                },
                {
                    "name": "■ 설정 업데이트 흐름",
                    "headers": ["단계", "함수 체인", "설명"],
                    "rows": [
                        (1, "subRoutineProcSaJob → procConfigUpdate",          "설정 변경 Job 수신"),
                        (2, "procConfigUpdate → updateConfigMem",               "메모리 설정 반영"),
                        (3, "updateConfigMem → updateAgent",                    "에이전트 전체 설정 갱신"),
                        (4, "updateAgent → updateSa / updateAm / updateModule", "SA/AM/모듈 설정 각각 갱신"),
                        (5, "updateSa → updateLogLevel",                        "로그 레벨 동적 변경"),
                    ],
                },
                {
                    "name": "■ 모듈 상태 전이",
                    "headers": ["상태", "전이 이벤트", "처리 함수"],
                    "rows": [
                        ("IDLE → RUNNING",   "Job 수신 → runModule",            "moduleStateUpdate(RUNNING)"),
                        ("RUNNING → FINISH", "모듈 gRPC procModuleFinish",      "moduleStateClear"),
                        ("RUNNING → DOWN",   "모듈 프로세스 비정상 종료",       "procModuleDownCase"),
                        ("RUNNING → KILL",   "procModuleKill Job",              "procTerminateModule"),
                        ("ANY → STOP",       "procStop / procStopAll Job",      "stopService / AllModuleStop"),
                    ],
                },
            ],
        },
        "deps": {
            "title": "SA 외부 의존성 및 연동",
            "sections": [
                {
                    "name": "■ 외부 서비스 연동",
                    "headers": ["서비스", "프로토콜", "방향", "주요 함수", "설명"],
                    "rows": [
                        ("AM (Agent Manager)", "HTTP REST", "SA → AM",  "sendPostRequest, sendDeleteRequest", "Job 폴링, 결과 전송, 로그 전송"),
                        ("각 모듈 (AA/AD/AF)", "gRPC",      "모듈 → SA","HeartBeat, CommonProcess",           "모듈 상태 보고, 완료 통보"),
                        ("GS (Gather Service)","내부 채널", "SA → GS",  "sendGsJob, RoutineGatherService",    "스캔 스케줄 실행"),
                        ("BoltDB (로컬)",       "파일 I/O",  "SA ↔ DB",  "OpenFileDb, InsertJob, DeleteJob",   "Job 영속화"),
                    ],
                },
                {
                    "name": "■ 내부 패키지 의존성",
                    "headers": ["패키지", "의존 패키지", "주요 사용처"],
                    "rows": [
                        ("moduleManager",        "internal/application/config, internal/common/...", "설정 조회, 디렉토리 탐색, 프로세스 실행"),
                        ("saJobManager",         "define, messageSender, moduleManager",             "Job 타입별 처리 위임"),
                        ("messageSender",        "internal/application/config, define",              "HTTP 통신, 설정 참조"),
                        ("grpcServer",           "internal/common/gear, proto",                      "gRPC 서비스 구현"),
                        ("dbManager",            "bolt (BoltDB), encoding/gob",                      "로컬 파일 DB 관리"),
                        ("logCollectionManager", "internal/common/file, define",                     "파일 읽기, 로그 상태 관리"),
                        ("resourceManager",      "internal/common/..., global",                      "시스템 리소스 측정, 알람 채널"),
                    ],
                },
                {
                    "name": "■ 공유 상태 (StRunMem)",
                    "headers": ["필드", "타입", "설명"],
                    "rows": [
                        ("SaJsonConfig",    "ConfigMem + Mutex",         "SA 설정 (sa.json) 메모리 캐시"),
                        ("AmJsonConfig",    "ConfigMem + Mutex",         "AM 설정 (am.json) 메모리 캐시"),
                        ("ModuleJsonConfig","ConfigMem + Mutex",         "모듈 설정 (module.json) 메모리 캐시"),
                        ("ModuleState",     "map[string]State + Mutex",  "실행 중 모듈 상태 맵"),
                    ],
                },
            ],
        },
    }


# ── main ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="SA 모듈 분석 xlsx 생성")
    parser.add_argument("output", nargs="?", default=None,
                        help="출력 파일 경로 (기본: /tmp/output.xlsx)")
    parser.add_argument("--data", default=None,
                        help="JSON 데이터 파일 경로")
    args = parser.parse_args()

    data = _default_data()
    if args.data and os.path.exists(args.data):
        with open(args.data, encoding="utf-8") as f:
            data = json.load(f)

    output_path = args.output or os.path.join(tempfile.gettempdir(), "output.xlsx")
    result = build(output_path, data)
    print(result)


if __name__ == "__main__":
    main()
